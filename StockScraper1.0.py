from lxml import html #import librarie
import requests #import librarie
import xlrd, xlwt #import librarie
from openpyxl import *
from tkinter import *
from tkinter import ttk

def gosearch(a):
    usersearch = a.title()
    print(usersearch)
    workbook = xlrd.open_workbook('CompanyFile.xlsx', on_demand=True)
    worksheet = workbook.sheet_by_index(0)
    i = 0
    g = 0
    counter = 0
    for i in range(worksheet.nrows):
        if usersearch in worksheet.cell(i,1).value:
            global ticker
            ticker = worksheet.cell_value(i,0)
            print("Your company is: ", worksheet.cell_value(i,1))
            search(ticker)
        else:
            counter += 1
        if counter == 3312:
            print("Company not found")

def directory():
    try:
        userchoice = int(input("\nPlease type the number appropriate to your wanted function, select 0 for all available functions"))
        if userchoice == 0:
            print("Select 1 to find up to date individual stock information")
            print("Select 2 to view your portfolio")
            usermenupick(int(input("Select a function")))
        else:
            usermenupick(userchoice)
    except:
        print("Invalid input")

def usermenupick(value):
    if value == 1:
        initialsearch = str(input("\nWhat is the name of the company?"))
        gosearch(initialsearch)
    elif value == 2:
        myportfolio()
        print("Your current portfolio is:", portfolio)
        userchoice = int(input(" Select 1 - To add a company to your portfolio\n "
                               "Select 2 - To remove a company from your portfolio IN DEVELOPMENT\n"
                               " Select 3 to return to main menu"))
        if userchoice == 1:
            addtoportfolio()

def search(stockticker): # definiton of seach function
    headers = requests.utils.default_headers()
    headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:20.0) Gecko/20100101 Firefox/20.0',
    })
    print('Stock ticker: ', stockticker)
    url = ("https://finance.yahoo.com/quote/%s" % stockticker)
    print(url)
    urlopen = requests.get(url, headers = headers)
    tree = html.fromstring(urlopen.content, "lxml")
    global price
    price = tree.xpath('//span[@class="Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)"]/text()')
    #description = tree.xpath('//p[@class="description__text"]/text()')
    global day_low
    global low_high
    day_low = tree.xpath('//span[@class="Trsdu(0.3s) "]/text()')
    low_high = tree.xpath('//td[@class="Ta(end) Fw(b) Lh(14px)"]/text()')
    #print('Description: ', description)
    print('Current stock price: $', price[0])
    print('Open stock price: $', day_low[1])
    print("Today's high and low prices: Low", low_high[0], 'High')
    myportfolio()
    if stockticker not in portfolio:
        PortfolioBtn = Button(tab1, text='Add to Portfolio', command=searchaddtoportfolio(ticker))
        PortfolioBtn.grid(column=2,row=4)
    global low_high_split
    low_high_split = low_high[0].split()
    print(low_high_split[2])
    return price, day_low, low_high_split
def intitialportfolio():
    global portfolio
    portfolio = []
    workbook = xlrd.open_workbook('CompanyFile.xlsx', on_demand=True)
    global row_count
    worksheet = workbook.sheet_by_index(1)
    global row_count
    row_count = worksheet.nrows
    if row_count != 0:
        for i in range(worksheet.nrows):
            cell = worksheet.cell(i, 0).value
            portfolio.append(cell)





def clicked():
    #pulls up info
    gosearch(txt.get())
    lbl1.configure(text = (txt.get()).title())
    currentP.configure(text = price)
    OpenP.configure(text = day_low[1])
    DayLow.configure(text = low_high_split[0])
    DayHigh.configure(text= low_high_split[2])




def searchaddtoportfolio(stockticker):
    wb = load_workbook(filename='CompanyFile.xlsx')
    sheet_ranges = wb['Sheet2']
    i = row_count
    print(row_count)
    if i == 1:
        sheet_ranges.cell(2, 1).value = stockticker
        wb.save('CompanyFile.xlsx')
    else:
        sheet_ranges.cell(i+1, 1).value = stockticker
        wb.save('CompanyFile.xlsx')

def addtoportfolio():
    searchtype = int(input("To search by ticker select 1 \n To search by company name select 2"))
    if searchtype == 1:
        companysearch = input("What is the company's ticker?").upper()
        search(companysearch)
        #validate ticker
    elif searchtype == 2:
        companysearch = input("What is the name of the company?")
        gosearch(companysearch)
        #validate company\

#def myportfolio():
 #   portfolio = []
###WATCHLIST###
def myportfolio():
    global portfolio
    portfolio = []
    workbook = xlrd.open_workbook('CompanyFile.xlsx', on_demand=True)
    global row_count
    worksheet = workbook.sheet_by_index(1)
    global row_count
    row_count = worksheet.nrows
    print('my portolio', row_count)
    if row_count != 0:
        for i in range(worksheet.nrows):
            cell = worksheet.cell(i,0).value
            portfolio.append(cell)
    portfolioBox = Listbox(tab2)
    portfolioBox.grid(column=0, row=0)
    for item in portfolio:
        portfolioBox.insert(END, item)


#while 1 == 1:
#    directory()



##    wallet = int(input("Enter your initial investment amount"))
##    def addcompany():
##        newcompany = input("Enter company name")
##        portfolio.append(name)
##        #search(name)
##        return portfolio
##    return portfolio
##def showportfolio():
##    #ALEC IS DOING THIS
##    #ART STUFF TO DISPLAY THE USERS PORTFOLIO
##    #USER SHOULD BE ABLE TO SELECT BEWTWEEN ALL THEIR PORFOLIOS
##
##def userportfolio(portfolio):
##    for name in portfolio:

intitialportfolio()

master = Tk()
# master.iconbitmap('Cucumber')
master.title('Cucumber')
tab_control = ttk.Notebook(master)
tab1 = ttk.Frame(tab_control)
tab2 = ttk.Frame(tab_control)
tab3 = ttk.Frame(tab_control)
tab_control.add(tab1, text='Search')
tab_control.add(tab2, text='Portfolio')
tab_control.add(tab3, text='Watchlist')
tab_control.grid(column=0,row=0)


lbl1 = Label(tab1, text='')
lbl1.grid(column=3, row=0)
currentP = Label(tab1, text='')
currentPLabel = Label(tab1, text=' Current Price:')
currentPLabel.grid(column=3, row=1)
currentP.grid(column=4, row=1)
OpenP = Label(tab1, text='')
OpenP.grid(column=4, row=2)
OpenPLabel = Label(tab1, text='Opening Price:')
OpenPLabel.grid(column=3, row=2)
DayLow = Label(tab1, text='')
DayLow.grid(column=4, row=3)
DayLowLabel = Label(tab1, text='Daily Low:')
DayLowLabel.grid(column=3, row=3)
DayHigh = Label(tab1, text='')
DayHigh.grid(column=4, row=4)
DayHighLabel = Label(tab1, text='Daily High:')
DayHighLabel.grid(column=3, row=4)
lbl = Label(tab1, text='Hello, Please enter the stock ticker or company name')
lbl.grid(column=2, row=1)
txt = Entry(tab1, width=10)
txt.grid(column=2, row=2)
lbl3 = Label(tab3, text='Feature still in development')
lbl3.grid(column=0,row=0)

btn = Button(tab1, text='Search', command=clicked)
btn.grid(column=2, row=3)
portfolioBox = Listbox(tab2)
portfolioBox.grid(column=0,row=0)
for item in portfolio:
    portfolioBox.insert(END, item)
master.mainloop()